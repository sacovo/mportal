""" This file is part of Sektionsportal.
    Sektionsportal is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    Sektionsportal is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with Sektionsportal.  If not, see <http://www.gnu.org/licenses/>.
"""
import datetime
import os
import tempfile
import time

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseBadRequest
from django.template import engines

from django.conf import settings
from django.contrib import messages
from django.contrib.admin.models import LogEntry, CHANGE
from django.contrib.contenttypes.models import ContentType
from django.contrib.auth.decorators import permission_required, login_required
from django.core.exceptions import PermissionDenied
from django.core import mail, serializers
from django.core.mail import send_mail as django_send_mail
from django.utils.translation import ugettext as _
from django.views.decorators.csrf import csrf_exempt
from django.contrib import admin
from django.contrib.auth import views as auth_views
from django.urls import reverse
from django.contrib.sites.shortcuts import get_current_site
from django.db.models import Q

from latex import build_pdf
from django.core.mail import EmailMultiAlternatives as EmailMessage
from traceback import print_exc
from openpyxl import load_workbook, Workbook
from twilio.rest import Client
from twilio import twiml
import reversion
from django.core.exceptions import ObjectDoesNotExist
import random
import string


from mportal.forms import EMailForm, ImportForm, SMSForm, LatexLetterForm, \
    ExportForm, member_fields, EMailFormMG, PortalFormForm, SMSAnswerForm
from mportal.models import Member, LetterTemplate, TemplateField,\
    SMSSender, MailGunEvent, MailGunMessage, CustomGroup,\
    PortalForm, PortalFormAnswer, ContactList, ImportMapping, Group,\
    SMS
from mportal.tasks import send_sms as send_sms_celery,\
        send_mail as send_mail_celery, send_mail_mg as send_mail_mg_celery


import hashlib as hl


def get_hash(s):
    return hl.sha512((s + settings.SECRET_KEY).encode('utf-8')).hexdigest()


def random_string(N=80):
    return ''.join(random.SystemRandom().choice(
        string.ascii_uppercase + string.digits) for _ in range(N))


@csrf_exempt
def mailgun_webhook(request):
    data = request.POST
    if 'message-id' not in data:
        return HttpResponse(status=204)
    try:
        int(data['message-id'])
    except Exception:
        return HttpResponse(status=204)
    message = MailGunMessage.objects.get(pk=data['message-id'])
    recipients = Member.objects.filter(email=data['recipient'])
    for recipient in recipients:
        first_open = recipient in message.unread()
        first_click = recipient in message.not_clicked()

        if data['event'] == 'opened' and first_open:
            recipient.activity_points += message.open_points
            recipient.read_mails.add(message)
            recipient.mail_activity = \
                recipient.read_mails.count()/recipient.sent_mails*100
            recipient.activate()
            recipient.save()

        elif data['event'] == 'clicked' and first_click:
            recipient.activity_points += message.click_points
            recipient.activate()

        elif data['event'] == 'delivered':
            recipient.sent_mails += 1
            recipient.mail_activity = \
                recipient.read_mails.count()/recipient.sent_mails*100
            recipient.save()

        event = MailGunEvent.objects.create(
            message=message,
            event=data['event'],
            recipient=recipient,
            country=data.get('country', ''),
            device_type=data.get('device-type', ''),
            client_type=data.get('client-type', ''),
            client_os=data.get('client-os', ''),
            tag=data.get('tag', ''),
            url=data.get('url', ''),
        )
    return HttpResponse(status=204)

def conversation_view(request, member, sender):
    groups = request.user.groups.all()
    sender = SMSSender.objects.get(pk=sender)
    member = Member.objects.get(pk=member)

    if not request.user.is_superuser:
        if sender.group not in request.user.groups.all():
            raise PermissionDenied

        qs = Member.objects.filter(Q(position__in=groups)|
                                   Q(imported_by=request.user))
        if not qs.filter(pk=member.pk).exists():
            if not member.sms_set.filter(twilio_account=sender).exists():
                raise PermissionDenied

    messages = member.sms_set.filter(twilio_account=sender)
    form = SMSAnswerForm()

    context = admin.site.each_context(request)
    context.update({
        'opts': SMS._meta,
        'title': _('Konversation'),
        'member': member,
        'sender': sender,
    })

    if request.method=='POST':
        form = SMSAnswerForm(request.POST)
        if form.is_valid():
            body = form.cleaned_data['content']
            auth = sender.auth
            sid = sender.sid
            user = sender.username
            client = Client(user, auth, user)
            message = client.messages.create(body=body, from_=sender.number,
                                             messaging_service_sid=sender.sid,
                                             to=member.phone_mobi,
                                             status_callback=settings.TWILIO_CALLBACK_URL)
            print(message.sid)
            sms = SMS.objects.create(
                    content=body,
                    member=member,
                    incoming=False,
                    twilio_account=sender,
                    is_mass=False,
                    twilio_sid=message.sid,
                    delivered=False,
                    )
            ct = ContentType.objects.get_for_model(SMS)
            LogEntry.objects.log_action(
                user_id=request.user.id,
                content_type_id=ct.pk,
                object_id=sms.pk,
                object_repr=sms.member.full_name,
                action_flag=CHANGE,
                change_message=_("SMS versandt: {}").format(sms.content)
                )
            form = SMSAnswerForm()

    context.update({'form':form})
    return render(request, 'mportal/conversation_view.html', context)

@csrf_exempt
def incoming_sms(request):

    body = request.POST['Body']
    sid = request.POST['MessagingServiceSid']
    to = request.POST['To']
    sender = SMSSender.objects.filter(sid=sid, number=to)[0]
    number = request.POST['From']
    location_fields = ['FromCity', 'FromState', 'FromZip', 'FromCountry']
    location_string = ','.join([
        request.POST[field] for field in location_fields
        if hasattr(request.POST, field)])

    auth = sender.auth
    sid = sender.sid
    user = sender.username

    client = Client(user, auth, sid)
    # Check if message existis in twilio
    try:
        if not client.messages.get(request.POST['MessageSid']):
            raise PermissionDenied
    except:
        raise PermissionDenied

    number_loc = client.lookups.phone_numbers(number).fetch().national_format
    number_loc = number_loc.replace(' ', '')

    qs = Member.objects.extra(where=["REPLACE(phone_mobi, ' ', '') = '{}'".format(number_loc)])
    if qs.exists():
        member = qs[0]
    else:
        member = Member.objects.create(phone_mobi=number_loc)
        member.save()

    sms = SMS.objects.create(
            content=body,
            member=member,
            incoming=True,
            twilio_account=sender,
            is_mass=False,
            location_string=location_string
        )

    return HttpResponse(status=204)

@csrf_exempt
def twilio_callback(request):
    if not request.POST['SmsStatus'] == 'delivered':
        return HttpResponse(status=204)
    sms = SMS.objects.get(twilio_sid=request.POST['MessageSid'])
    sms.delivered = True
    sms.save()
    return HttpResponse(status=204)

def open_chat(request, member):

    groups = request.user.groups.all()
    member = Member.objects.get(pk=member)

    if not request.user.is_superuser:

        qs = Member.objects.filter(Q(position__in=groups)|
                                   Q(imported_by=request.user))
        if not qs.filter(pk=member.pk).exists():
            raise PermissionDenied

    senders = SMSSender.objects.filter(group__in=groups)
    context = admin.site.each_context(request)
    context.update({
        'opts': SMSSender._meta,
        'title': _('Konversation'),
        'member': member.pk,
        'senders': senders,
    })
    return render(request, 'mportal/chose_twilio.html', context)


def chat_window(request, member, sender):
    groups = request.user.groups.all()
    member = Member.objects.get(pk=member)
    sender = SMSSender.objects.get(pk=sender)
    messages = member.sms_set.filter(twilio_account=sender).order_by('time')

    return render(request, 'mportal/chat_window.html', {'messages':messages})


@permission_required('mportal.send_sms')
def send_sms(request):
    form = SMSForm(user=request.user)
    recipients = (int(x) for x in request.session['recipients'].split(','))
    recipients = Member.objects.filter(pk__in=recipients)
    if request.method == 'POST':
        form = SMSForm(request.POST, user=request.user)
        if form.is_valid():
            content = form.cleaned_data['body']
            group = form.cleaned_data['group']
            send_sms_celery.delay(content, serializers.serialize(
                'json', recipients), group.pk)
            return redirect('admin:mportal_member_changelist')
    context = admin.site.each_context(request)
    context.update({
        'opts': Member._meta,
        'title': _('SMS verfassen'),
        'recipients': recipients,
        'form': form,
    })
    return render(request, 'mportal/send_sms.html', context)


@permission_required('mportal.change_member')
def send_mail(request, method=1):
    method = int(method)
    form = EMailForm()
    recipients = (int(x) for x in request.session['recipients'].split(','))
    recipients = Member.objects.filter(pk__in=recipients)
    r = 'admin:mportal_member_changelist'

    if request.method == 'POST':
        form = EMailForm(request.POST, request.FILES)
        if form.is_valid():
            attachements = []
            data = form.cleaned_data
            for x in range(1, 4):
                if data['attach_' + str(x)]:
                    f = data['attach_' + str(x)]
                    name = f.name
                    t = open('/tmp/attach_'+str(x), 'wb')
                    t.write(f.read())
                    t.close()
                    attachements.append({'name': name,
                                         'path': '/tmp/attach_'+str(x), })
                    del data['attach_' + str(x)]
            form.cleaned_data['template'] = form.cleaned_data['template'].pk
            send_mail_celery.delay(
                    form.cleaned_data,
                    serializers.serialize('json', recipients),
                    attachements
                )
            return redirect(r)
    context = admin.site.each_context(request)
    context.update({
        'opts': Member._meta,
        'title': _('E-Mail'),
        'recipients': recipients,
        'form': form,
    })
    return render(request, 'mportal/send_mail.html', context)


def list_is(l, ix):
    ln = []
    for i in ix:
        ln.append(l[i])
        return ln


def handle_import_file(excel, status, user, contact_list,
                       import_mapping, positions=[],
                       do_reversion=False, delete=True):
    start_time = time.time()
    default_groups = [Group.objects.get(pk=p) for p in positions]
    contact_list = ContactList.objects.get(pk=contact_list)
    import_mapping = ImportMapping.objects.get(pk=import_mapping)
    with tempfile.NamedTemporaryFile('w+b', suffix='.xlsx') as import_file:
        for chunk in excel.chunks():
            import_file.write(chunk)
        import_file.flush()
        ws = load_workbook(import_file.name, True,
                           guess_types=True).worksheets[0]

        simple_colnames = [c.column_name for c in
                           import_mapping.fields.exclude(var_name='position').
                           exclude(var_name__startswith="birthday")]
        simple_col_indices = {cell.value: cell.column - 1 for cell in ws[1]
                              if cell.value in simple_colnames}

        if import_mapping.birthday_multiple():
            birthday_colnames = import_mapping.birthday_col_names()
            birthday_col_indices =\
                sorted([cell.column - 1 for cell in ws[1]
                       if cell.value in birthday_colnames], reverse=True)
        elif import_mapping.birthday_single():
            birthday_colnames = [import_mapping.fields.get(
                var_name="birthday").column_name]
            birthday_col_indices =\
                sorted([cell.column - 1 for cell in ws[1]
                       if cell.value in birthday_colnames], reverse=True)

        position_colnames = [
            p.column_name for p in import_mapping.fields.filter(var_name="position")]
        position_col_indices = [cell.column - 1 for cell in ws[1]
                                if cell.value in position_colnames]

        om_members = set()
        tuple_cls = Member.as_named_tuple()
        first = True

        def empty_line(m_dict):
            for key, value in m_dict.items():
                if not (key in ('status', 'position', 'om_number')) and value:
                    return False
            return True
        increment = 0
        mapped_fields = list(import_mapping.fields.exclude(var_name="position")
                             .exclude(var_name__startswith="birthday"))
        birthday_single = import_mapping.birthday_single()
        birthday_multiple = import_mapping.birthday_multiple()

        for row in ws.rows:
            if first:
                first = False
                continue

            member_dict = dict(
                position=frozenset(
                    {row[c].value for c in position_col_indices}),
                status=status,
                contact_list=contact_list
            )
            if (None in member_dict['position'] or
                    len(member_dict['position']) == 0) and default_groups:
                member_dict = dict(
                    position=frozenset({p.name for p in default_groups}),
                    status=status,
                    contact_list=contact_list
                )
            member_dict.update({
                'city': '', 'country': '', 'phone_priv': '',
                'phone_busi': '', 'phone_mobi': '', 'gender': '',
                'letter_opening': '', 'om_number': '', 'first_name': '',
                'last_name': '', 'street': '', 'plz': '', 'country': '',
                'birthday': None,
            })
            for f in mapped_fields:
                member_dict[f.var_name] = str(
                    row[simple_col_indices[f.column_name]].value) or ''

            if birthday_multiple:
                bd_string = ".".join([str(row[c].value or 0)
                                      for c in birthday_col_indices])
                member_dict['birthday'] = datetime.date.fromtimestamp(
                    time.mktime(time.strptime(bd_string, '%d.%m.%Y'))) if not bd_string.startswith(
                        "0.") and not bd_string.endswith(".0") else None
            elif birthday_single:
                member_dict['birthday'] = row[birthday_col_indices[0]
                                              ].value or None
            if member_dict['om_number'] == '':
                increment += 1
                member_dict['om_number'] = increment
            member_dict['om_number'] = str(member_dict['om_number'])
            if empty_line(member_dict):
                break
            print(member_dict)
            om_members.add(tuple_cls(**member_dict))

    sektionsportal_members = Member.get_all_as_named_tuple(contact_list)

    new_members = set()
    updated_members = set()
    deleted_members = set()
    for m in om_members ^ sektionsportal_members:
        if m.om_number not in {x.om_number for x in sektionsportal_members}:
            new_members.add(m)
        elif m.om_number in {x.om_number for x in om_members}:
            # Just take the one from the import for the update
            if m in om_members:
                updated_members.add(m)
        elif m.status == status and delete:
            deleted_members.add(m)

    def member_to_dict(m):
        return dict(om_number=m.om_number or 0,
                    anrede="",
                    first_name=m.first_name,
                    last_name=m.last_name,
                    street=m.street,
                    plz=m.plz,
                    city=m.city,
                    country=m.country,
                    phone_priv=m.phone_priv,
                    phone_busi=m.phone_busi,
                    phone_mobi=m.phone_mobi,
                    email=m.email,
                    gender=m.gender,
                    letter_opening=m.letter_opening,
                    birthday=m.birthday,
                    status=m.status,
                    contact_list=contact_list,
                    imported_by=user
                    )

    def get_groups(m):
        groups = []
        for g in m.position:
            if g:
                group, exists = Group.objects.get_or_create(name=g)
                if not exists:
                    group.save()
                groups.append(group)
        return groups

    for m in updated_members:
        existing = Member.objects.filter(
            om_number=m.om_number, contact_list=contact_list)
        if do_reversion:
            with reversion.create_revision():
                existing = existing[0]
                existing.__dict__.update(member_to_dict(m))
                existing.position.add(*get_groups(m))
                existing.position.add(*default_groups)
                existing.save()
                reversion.set_user(user)
                reversion.set_comment("Updated by excel")
        else:
            existing = existing[0]
            existing.__dict__.update(member_to_dict(m))
            existing.position.add(*get_groups(m))
            existing.position.add(*default_groups)
            existing.save()


    for m in new_members:
        if do_reversion:
            with reversion.create_revision():
                new_m = Member(
                    **member_to_dict(m)
                )
                new_m.save()
                new_m.position.set(get_groups(m))
                new_m.save()
                reversion.set_user(user)
                reversion.set_comment("First excel import")
        else:
            new_m = Member(
                **member_to_dict(m)
            )
            new_m.save()
            new_m.position.set(get_groups(m))
            new_m.save()


    for m in deleted_members:
        if do_reversion:
            with reversion.create_revision():
                Member.objects.filter(
                    om_number=m.om_number, contact_list=contact_list).delete()
                reversion.set_user(user)
                reversion.set_comment("Deleted with excel import")
        else:
            Member.objects.filter(om_number=m.om_number,
                                  contact_list=contact_list).delete()

    return len(new_members), len(deleted_members), time.time() - start_time


@permission_required('mportal.add_member')
def _import(request):
    form = ImportForm()
    form.fields['import_mapping'].choices = [
        (i.pk, i.name) for i in ImportMapping.objects.all()]
    form.fields['positions'].choices = [
        (g.pk, g.name) for g in request.user.groups.all()]
    form.fields['contact_list'].choices = [
        (c.pk, c.name) for c in ContactList.objects.filter(users=request.user)
    ]
    if request.user.is_superuser:
        form.fields['positions'].choices = [
            (g.pk, g.name) for g in Group.objects.all()]
        form.fields['contact_list'].choices = [
            (c.pk, c.name) for c in ContactList.objects.all()
        ]

    if request.method == 'POST':
        form = ImportForm(request.POST, request.FILES)
        form.fields['import_mapping'].choices = [
            (i.pk, i.name) for i in ImportMapping.objects.all()]
        form.fields['positions'].choices = [
            (g.pk, g.name) for g in request.user.groups.all()]
        form.fields['contact_list'].choices = [
            (c.pk, c.name) for c in ContactList.objects.filter(users=request.user)
        ]
        if request.user.is_superuser:
            form.fields['positions'].choices = [
                (g.pk, g.name) for g in Group.objects.all()]
            form.fields['contact_list'].choices = [
                (c.pk, c.name) for c in ContactList.objects.all()
            ]
        if form.is_valid():
            positions = form.cleaned_data['positions']
            n, d, t = handle_import_file(request.FILES['excel'], form.cleaned_data['status'],
                                         request.user, form.cleaned_data['contact_list'],
                                         form.cleaned_data['import_mapping'], positions, form.cleaned_data['reversion'],
                                         delete=form.cleaned_data['delete'])
            messages.success(request, _(
                "Update erfolgreich, {n} hinzugefügt und {d} gelöscht. (Dauer: {t:.4f} s)".format(n=n, d=d, t=t)
                ))
            return redirect('admin:mportal_member_changelist')

    context = admin.site.each_context(request)
    context.update({
        'opts': Member._meta,
        'title': "Importieren",
        'form': form,
    })
    return render(request, 'mportal/import.html', context)


def chose_template(request, method):
    templates = LetterTemplate.objects.all()
    context = admin.site.each_context(request)
    context.update({
        'opts': Member._meta,
        'title': _('Vorlage wählen'),
        'templates': templates,
        'method': int(method),
    })
    return render(request, 'mportal/chose_template.html', context)


def save_file_to(f, path):
    with open(path, 'wb+') as dest:
        for chunk in f.chunks():
            dest.write(chunk)


def export_excel(request):
    recipients = (int(x) for x in request.session['recipients'].split(','))
    recipients = Member.objects.filter(pk__in=recipients)
    r = 'admin:mportal_member_changelist'
    opts = Member._meta
    form = ExportForm(request.GET)
    if form.is_valid():
        columns = form.cleaned_data['columns']
        column_titles = dict(member_fields())
        wb = Workbook()
        ws = wb.active
        ws.title = _("Mitglieder")
        for i, field in enumerate(columns, start=1):
            r = ws.cell(row=1, column=i, value=column_titles[field])
            for i, member in enumerate(recipients, start=2):
                for j, field in enumerate(columns, start=1):
                    r = ws.cell(row=i, column=j, value=getattr(member, field))

        filename = 'Export_{:%Y%m%d}.xlsx'.format(datetime.datetime.now())
        response = HttpResponse(content_type='application/xlsx')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
        wb.save(response)

        ct = ContentType.objects.get_for_model(recipients.model)
        for obj in recipients:
            LogEntry.objects.log_action(
                user_id=request.user.id,
                content_type_id=ct.pk,
                object_id=obj.pk,
                object_repr=obj.full_name,
                action_flag=CHANGE,
                change_message="Excel exportiert: {}".format(columns)
                )

        return response
    context = admin.site.each_context(request)
    context.update({
        'opts': Member._meta,
        'title': _('Felder wählen'),
        'form': ExportForm(),
    })
    return render(request, 'mportal/select_fields.html', context)


def render_letter(request, template):
    recipients = (int(x) for x in request.session['recipients'].split(','))
    recipients = Member.objects.filter(pk__in=recipients)
    template = LetterTemplate.objects.get(pk=template)
    form = LatexLetterForm(template=template)
    r = 'admin:mportal_member_changelist'
    opts = Member._meta
    if request.method == 'POST':
        form = LatexLetterForm(request.POST, request.FILES, template=template)
        if form.is_valid():
            string = template.template.read().decode()
            t = engines['django'].from_string(string)
            context = form.cleaned_data.copy()
            context['recipients'] = recipients
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="letter.pdf"'
            src = t.render(context=context, request=request)
            latex_root = os.path.join(settings.MEDIA_ROOT, 'latex')
            for f in template.fields.all():
                if f.input_type == 'forms.FileField' and\
                   request.FILES.get(f.var_name, None):
                    fi = request.FILES[f.var_name]
                    save_file_to(fi, os.path.join(latex_root, f.var_name))
            pdf = build_pdf(src, texinputs=[latex_root])
            response.write(bytes(pdf))
            return response
    context = admin.site.each_context(request)
    context.update({
        'opts': opts,
        'title': _('Brief generieren'),
        'recipients': recipients,
        'form': form,
    })
    return render(request, 'mportal/render_letter.html', context)


@permission_required('mportal.change_member')
def send_mail_mg(request):
    form = EMailFormMG(user=request.user)
    recipients = (int(x) for x in request.session['recipients'].split(','))
    recipients = Member.objects.filter(pk__in=recipients)
    r = 'admin:mportal_member_changelist'
    opts = Member._meta

    if request.method == 'POST':
        form = EMailFormMG(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            attachements = []
            for x in form.cleaned_data.pop('attachement'):
                name = x.name
                t = open('/tmp/attach_'+str(x), 'wb')
                t.write(x.read())
                t.close()
                attachements.append({'name': name,
                                     'path': '/tmp/attach_'+str(x), })
            group = form.cleaned_data.pop('group')
            form.cleaned_data['template'] = form.cleaned_data['template'].pk
            send_mail_mg_celery.delay(form.cleaned_data,
                                      serializers.serialize('json', recipients), attachements, group.pk)
            return redirect(r)

    context = admin.site.each_context(request)
    context.update({
        'opts': opts,
        'title': _('E-Mail (Mailgun)'),
        'recipients': recipients,
        'form': form,
    })
    return render(request, 'mportal/send_mail.html', context)


def send_confirmation_link(answer, request):
    form = answer.portal_form
    s = random_string()
    link = 'https://' + get_current_site(request).domain + \
        reverse("form-confirmation", kwargs={'conf': str(s)})
    django_send_mail(
        "{form.name} bestätigen".format(form=form),
        form.confirmation_text.format(member=answer.member, form=form,
                                      confirmation_link=link),
        "{form.contact_name} <{form.contact_mail}>".format(form=form),
        [answer.member.email],
        True
    )
    answer.confirmation_string = get_hash(s)
    answer.save()


def portalform_view(request, slug):
    portalform = get_object_or_404(PortalForm, slug=slug)
    form = PortalFormForm(portalform=portalform)
    context = {
        'form': form,
        'title': portalform.name,
        'description': portalform.description,
        'portalform': portalform,
    }
    if request.user in portalform.users.all() or request.user.is_superuser:
        context.update({'can_edit': True})
    if request.method == 'POST':
        form = PortalFormForm(request.POST, request.FILES,
                              portalform=portalform)
        if form.is_valid():
            data = form.cleaned_data
            e_mail = data[portalform.email_field_name]
            try:
                m = Member.objects.filter(position__in=portalform.allowed_groups.all())\
                    .distinct().get(email_iexact=e_mail)
                if not PortalFormAnswer.objects.filter(portal_form=portalform, member=m).exists():
                    answer = PortalFormAnswer.objects.create(
                        portal_form=portalform,
                        member=m,
                        confirmation_string=random_string(),
                        confirmed=False
                    )
                    for field in portalform.fields.all():
                        answer.answers.create(
                            form_field=field, value=data[field.var_name])
                    send_confirmation_link(answer, request)
                    return render(request, 'mportal/link_sent.html')
                else:
                    form.add_error(portalform.email_field_name,
                                   'Du hast das Formular bereits ausgefüllt!')
                    answer = PortalFormAnswer.objects.filter(
                        portal_form=portalform, member=m)[0]
                    context.update({'answer': answer})

            except ObjectDoesNotExist:
                form.add_error(portalform.email_field_name,
                               'E-Mail Addresse nicht gefunden!')

    context.update({'form': form})
    return render(request, 'mportal/form_view.html', context)


def confirm_view(request, conf):
    h = get_hash(conf)
    answer = get_object_or_404(PortalFormAnswer, confirmation_string=h)
    portalform = answer.portal_form
    if request.method == 'POST':
        if portalform.editable:
            form = PortalFormForm(
                request.POST, request.FILES, portalform=portalform, answer=answer)
            if form.is_valid():
                data = form.cleaned_data
                for field in portalform.fields.exclude(var_name=portalform.email_field_name):
                    if answer.answers.filter(form_field=field).exists():
                        answer.answers.filter(form_field=field).update(
                            value=data[field.var_name])
                    else:
                        answer.answers.create(
                            form_field=field, value=data[field.var_name])
                if data['confirm_answer']:
                    answer.confirmed = True
                else:
                    answer.confirmed = False
                if answer.confirmed and portalform.destination_group:
                    answer.member.position.add(portalform.destination_group)
                elif portalform.destination_group:
                    answer.member.position.remove(portalform.destination_group)
                answer.save()
            return render(request, 'mportal/confirmation.html', dict(
                portalform=portalform,
                answer=answer,
                title=portalform.name,
                message=True,
                form=form,
            ))
        else:
            answer.confirmed = not answer.confirmed
            if answer.confirmed and portalform.destination_group:
                answer.member.position.add(portalform.destination_group)
            elif portalform.destination_group:
                answer.member.position.remove(portalform.destination_group)
            answer.save()
            return render(request, 'mportal/confirmation.html', dict(
                portalform=portalform,
                answer=answer,
                title=portalform.name,
                message=True,
            ))
    if portalform.editable:
        initial = dict()
        for field in portalform.fields.all():
            initial[field.var_name] = answer.answers.get(
                form_field=field).value
        form = PortalFormForm(portalform=portalform,
                              initial=initial, answer=answer)
        return render(request, 'mportal/confirmation.html', dict(
            portalform=portalform,
            title=portalform.name,
            answer=answer,
            form=form,
        ))

    return render(request, 'mportal/confirmation.html', dict(
        portalform=portalform,
        title=portalform.name,
        answer=answer
    ))


def send_form_link(request):
    if request.POST:
        answer = get_object_or_404(
            PortalFormAnswer, pk=int(request.POST['pk']))
        send_confirmation_link(answer, request)
    return render(request, 'mportal/link_sent.html')


@login_required
def form_result_view(request, slug):
    form = get_object_or_404(PortalForm, slug=slug)
    if not request.user in form.users.all():
        return PermissionDenied
    answers = form.portalformanswer_set.filter(confirmed=True)
    field_names = list(form.fields.all())
    member_fields = list(form.memberfield_set.all())
    answer_list = []
    for answer in answers:
        tmp = []
        for mf in member_fields:
            value = getattr(answer.member, mf.var_name, '---')
            tmp.append(value)
        for fn in field_names:
            if answer.answers.filter(form_field=fn).exists():
                value = answer.answers.get(form_field=fn)
                tmp.append(value.value)
            else:
                tmp.append('---')
        answer_list.append(tmp)
    context = admin.site.each_context(request)
    context.update({
        'title': form.name + _(" Resultate"),
        'form': form,
        'answers': answer_list,
        'fields': member_fields + field_names,
        'opts': PortalForm._meta,
    })
    return render(request, "mportal/form_result.html", context)


@login_required
def mailgun_report(request, pk):
    selected = get_object_or_404(MailGunMessage, pk=pk)
    if selected.group in request.user.groups.all():
        context = admin.site.each_context(request)
        context.update({
            'opts': MailGunMessage._meta,
            'message': selected,
        })
        return render(request, 'mportal/mailgun_report.html', context)
    raise PermissionDenied
